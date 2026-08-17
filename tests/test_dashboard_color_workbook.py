"""Focused tests for the colleague-facing dashboard colour workbook."""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill

from codebase.dashboard_color_config import average_oklab
from scripts.manage_dashboard_colors import (
    REQUIRED_EXTERNAL_FUELS,
    REQUIRED_EXTERNAL_SECTORS,
    export_color_workbook,
    import_color_workbook,
    resolve_json_synced_colors,
    synchronize_dashboard_colors,
)


def _sample_config(path: Path) -> None:
    path.write_text(
        json.dumps(
            {
                "plotting": {
                    "product": {"Coal": "#111111"},
                    "flow": {},
                    "capacity": {},
                    "series": {
                        "ESTO Historical": "#0072B2",
                        "LEAP Reference": "#D55E00",
                        "LEAP Target": "#D55E00",
                        "9th Reference": "#009E73",
                        "9th Target": "#009E73",
                    },
                },
                "product": {
                    "01": "#0D0D0D",
                    "01.02": "#A6A6A6",
                    "01.03": "#C4C4C4",
                    "01.04": "#7D7D7D",
                    "17": "#FFD757",
                    "22": "#00B9CC",
                    "23": "#F37021",
                },
                "flow": {
                    "01": "#2F855A",
                    "14": "#AA0000",
                    "18.01": "#362B4F",
                    "18.02": "#8A5DA7",
                    "18.03": "#0F266B",
                },
            }
        ),
        encoding="utf-8",
    )


def _sample_common_rows(path: Path) -> None:
    path.write_text(
        "comparison_scope,common_product_code,common_product_label,component_product_code,common_flow_code,common_flow_label,component_flow_code\n"
        "esto_leap_ninth,01.02-01.04,01.02-01.04 Coal,01.02,18.01-18.03,18.01-18.03 Electricity output in GWh,18.01\n"
        "esto_leap_ninth,01.02-01.04,01.02-01.04 Coal,01.03,18.01-18.03,18.01-18.03 Electricity output in GWh,18.02\n"
        "esto_leap_ninth,01.02-01.04,01.02-01.04 Coal,01.04,18.01-18.03,18.01-18.03 Electricity output in GWh,18.03\n",
        encoding="utf-8",
    )


def _sample_external_colors(path: Path) -> None:
    rows = [
        {
            "code": code,
            "label": label,
            "hex": "#101010" if code == "01" else f"#{index:06X}",
        }
        for index, (code, label) in enumerate(REQUIRED_EXTERNAL_FUELS.items(), start=1)
    ]
    sectors = [
        {"code": code, "label": label, "hex": f"#{index + 100:06X}"}
        for index, (code, label) in enumerate(REQUIRED_EXTERNAL_SECTORS.items(), start=1)
    ]
    path.write_text(
        json.dumps({"fuels": {"standard": rows}, "sectors": {"list": sectors}}),
        encoding="utf-8",
    )


def _row_for_key(sheet: object, key: str) -> int:
    return next(row for row in range(2, sheet.max_row + 1) if sheet.cell(row=row, column=1).value == key)


def _metadata_row_for_key(sheet: object, sheet_name: str, key: str) -> int:
    return next(
        row
        for row in range(8, sheet.max_row + 1)
        if sheet.cell(row=row, column=1).value == sheet_name
        and sheet.cell(row=row, column=2).value == key
    )


def test_export_and_import_accept_typed_and_fill_changes(tmp_path: Path) -> None:
    config_path = tmp_path / "code_colors.json"
    custom_path = tmp_path / "code_colors_custom.json"
    workbook_path = tmp_path / "colors.xlsx"
    common_rows_path = tmp_path / "common_rows.csv"
    _sample_config(config_path)
    _sample_common_rows(common_rows_path)

    export_color_workbook(output_path=workbook_path, code_colors_path=config_path, common_rows_path=common_rows_path, external_colors_path=None)
    workbook = load_workbook(workbook_path)
    assert workbook["_metadata"].sheet_state == "veryHidden"
    assert workbook["_metadata"]["B1"].value == "7"
    assert [workbook["_metadata"].cell(row=7, column=column).value for column in range(1, 9)] == [
        "sheet",
        "key",
        "category",
        "current_color",
        "sync_with_json",
        "exists_in_json",
        "color_components",
        "missing_color_components",
    ]
    assert workbook["Products"].column_dimensions["A"].hidden is True
    assert workbook["Products"]["B2"].value == "01 Coal"
    assert workbook["Products"]["C2"].value == "#0D0D0D"
    assert workbook["Products"]["C2"].fill.fgColor.rgb.endswith("0D0D0D")
    assert workbook["Products"]["D1"].value == "SYNC_WITH_JSON"
    assert workbook["Products"]["D2"].value == "FALSE"
    assert workbook["Products"]["E1"].value == "EXISTS_IN_JSON"
    assert workbook["Products"]["E2"].value == "FALSE"
    assert workbook["Products"].auto_filter.ref is None
    assert list(workbook["Products"].tables) == ["ProductsColours"]
    assert workbook["Flows"].auto_filter.ref is None
    assert list(workbook["Flows"].tables) == ["FlowsColours"]
    assert workbook["Other categories"].auto_filter.ref is None
    assert list(workbook["Other categories"].tables) == ["OthercategoriesColours"]
    assert workbook["Other categories"]["D1"].value is None
    assert workbook["Other categories"]["E1"].value is None
    assert not any(
        str(workbook["Products"].cell(row=row, column=1).value or "").startswith("common::")
        for row in range(2, workbook["Products"].max_row + 1)
    )
    assert "22" not in {
        str(workbook["Products"].cell(row=row, column=1).value or "")
        for row in range(2, workbook["Products"].max_row + 1)
    }
    product_row = _row_for_key(workbook["Products"], "01.02")
    flow_row = _row_for_key(workbook["Flows"], "01")
    other_row = _row_for_key(workbook["Other categories"], "LEAP Target")
    workbook["Products"].cell(product_row, 3).value = "#123456"
    workbook["Flows"].cell(flow_row, 3).fill = PatternFill("solid", fgColor="ABCDEF")
    workbook["Other categories"].cell(other_row, 3).value = "#654321"
    workbook.save(workbook_path)

    import_color_workbook(workbook_path, config_path, custom_path, common_rows_path, external_colors_path=None)
    custom = json.loads(custom_path.read_text(encoding="utf-8"))
    production = json.loads(config_path.read_text(encoding="utf-8"))
    assert custom["product"]["01.02"] == "#123456"
    assert custom["flow"]["01"] == "#ABCDEF"
    assert production["product"]["01.02"] == "#123456"
    assert production["product"]["22"] == "#00B9CC"
    assert custom["product"]["22"] == "#00B9CC"
    assert production["flow"]["01"] == "#ABCDEF"
    assert custom["plotting"]["series"]["LEAP Target"] == "#654321"
    assert production["plotting"]["series"]["LEAP Target"] == "#654321"
    assert production["common"]["product"]["01.02-01.04"] == average_oklab(["#123456", "#C4C4C4", "#7D7D7D"])
    assert custom["common_overrides"] == {"product": {}, "flow": {}}
    assert set(workbook.sheetnames) == {"START HERE", "Products", "Flows", "Other categories", "_metadata"}


def test_import_rejects_conflicting_typed_and_fill_changes(tmp_path: Path) -> None:
    config_path = tmp_path / "code_colors.json"
    workbook_path = tmp_path / "colors.xlsx"
    common_rows_path = tmp_path / "common_rows.csv"
    _sample_config(config_path)
    _sample_common_rows(common_rows_path)
    export_color_workbook(output_path=workbook_path, code_colors_path=config_path, common_rows_path=common_rows_path, external_colors_path=None)
    workbook = load_workbook(workbook_path)
    workbook["Products"]["C2"] = "#123456"
    workbook["Products"]["C2"].fill = PatternFill("solid", fgColor="654321")
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="disagree"):
        import_color_workbook(workbook_path, config_path, tmp_path / "custom.json", common_rows_path, external_colors_path=None)


def test_import_accepts_excel_theme_fill(tmp_path: Path) -> None:
    config_path = tmp_path / "code_colors.json"
    workbook_path = tmp_path / "colors.xlsx"
    custom_path = tmp_path / "custom.json"
    common_rows_path = tmp_path / "common_rows.csv"
    _sample_config(config_path)
    _sample_common_rows(common_rows_path)
    export_color_workbook(output_path=workbook_path, code_colors_path=config_path, common_rows_path=common_rows_path, external_colors_path=None)
    workbook = load_workbook(workbook_path)
    workbook["Products"]["C2"].fill = PatternFill("solid", fgColor=Color(theme=4))
    workbook.save(workbook_path)

    import_color_workbook(workbook_path, config_path, custom_path, common_rows_path, external_colors_path=None)
    custom = json.loads(custom_path.read_text(encoding="utf-8"))
    assert custom["product"]["01"] == "#4F81BD"


def test_import_rejects_changed_category_label(tmp_path: Path) -> None:
    config_path = tmp_path / "code_colors.json"
    workbook_path = tmp_path / "colors.xlsx"
    custom_path = tmp_path / "custom.json"
    common_rows_path = tmp_path / "common_rows.csv"
    _sample_config(config_path)
    _sample_common_rows(common_rows_path)
    export_color_workbook(output_path=workbook_path, code_colors_path=config_path, common_rows_path=common_rows_path, external_colors_path=None)
    workbook = load_workbook(workbook_path)
    product_row = _row_for_key(workbook["Products"], "01.02")
    workbook["Products"].cell(product_row, 2).value = "Overwritten category"
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="category was changed"):
        import_color_workbook(workbook_path, config_path, custom_path, common_rows_path, external_colors_path=None)


def test_import_rejects_stale_workbook_before_overwriting_new_config(tmp_path: Path) -> None:
    config_path = tmp_path / "code_colors.json"
    workbook_path = tmp_path / "colors.xlsx"
    common_rows_path = tmp_path / "common_rows.csv"
    _sample_config(config_path)
    _sample_common_rows(common_rows_path)
    export_color_workbook(output_path=workbook_path, code_colors_path=config_path, common_rows_path=common_rows_path, external_colors_path=None)

    changed_config = json.loads(config_path.read_text(encoding="utf-8"))
    changed_config["product"]["01.02"] = "#FFFFFF"
    config_path.write_text(json.dumps(changed_config), encoding="utf-8")

    with pytest.raises(ValueError, match="changed after this workbook was exported"):
        import_color_workbook(workbook_path, config_path, tmp_path / "custom.json", common_rows_path, external_colors_path=None)


def test_import_rejects_deleted_category_row(tmp_path: Path) -> None:
    config_path = tmp_path / "code_colors.json"
    workbook_path = tmp_path / "colors.xlsx"
    common_rows_path = tmp_path / "common_rows.csv"
    _sample_config(config_path)
    _sample_common_rows(common_rows_path)
    export_color_workbook(output_path=workbook_path, code_colors_path=config_path, common_rows_path=common_rows_path, external_colors_path=None)
    workbook = load_workbook(workbook_path)
    product_row = _row_for_key(workbook["Products"], "01.02")
    workbook["Products"].delete_rows(product_row)
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="category rows changed"):
        import_color_workbook(workbook_path, config_path, tmp_path / "custom.json", common_rows_path, external_colors_path=None)


def test_import_rejects_edited_exists_in_json_indicator(tmp_path: Path) -> None:
    config_path = tmp_path / "code_colors.json"
    workbook_path = tmp_path / "colors.xlsx"
    common_rows_path = tmp_path / "common_rows.csv"
    external_path = tmp_path / "colors.json"
    _sample_config(config_path)
    _sample_common_rows(common_rows_path)
    _sample_external_colors(external_path)
    export_color_workbook(workbook_path, config_path, common_rows_path, external_path)
    workbook = load_workbook(workbook_path)
    coal_row = _row_for_key(workbook["Products"], "01")
    workbook["Products"].cell(coal_row, 5).value = "FALSE"
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="EXISTS_IN_JSON is automatic"):
        import_color_workbook(
            workbook_path,
            config_path,
            tmp_path / "custom.json",
            common_rows_path,
            external_colors_path=external_path,
        )


def test_export_syncs_from_detailed_external_json(tmp_path: Path) -> None:
    config_path = tmp_path / "code_colors.json"
    custom_path = tmp_path / "code_colors_custom.json"
    workbook_path = tmp_path / "colors.xlsx"
    common_rows_path = tmp_path / "common_rows.csv"
    external_path = tmp_path / "colors.json"
    _sample_config(config_path)
    _sample_common_rows(common_rows_path)
    _sample_external_colors(external_path)

    export_color_workbook(
        output_path=workbook_path,
        code_colors_path=config_path,
        common_rows_path=common_rows_path,
        external_colors_path=external_path,
    )
    workbook = load_workbook(workbook_path)
    coal_row = _row_for_key(workbook["Products"], "01")
    assert workbook["Products"].cell(coal_row, 3).value == "#101010"
    assert workbook["Products"].cell(coal_row, 3).fill.fgColor.rgb.endswith("101010")
    assert workbook["Products"].cell(coal_row, 4).value == "TRUE"
    assert workbook["Products"].cell(coal_row, 5).value == "TRUE"
    coal_metadata_row = _metadata_row_for_key(workbook["_metadata"], "Products", "01")
    assert workbook["_metadata"].cell(coal_metadata_row, 7).value is None
    assert workbook["_metadata"].cell(coal_metadata_row, 8).value is None
    unmapped_product_row = _row_for_key(workbook["Products"], "01.02")
    assert workbook["Products"].cell(unmapped_product_row, 4).value == "TRUE"
    assert workbook["Products"].cell(unmapped_product_row, 5).value == "FALSE"
    industry_row = _row_for_key(workbook["Flows"], "14")
    assert workbook["Flows"].cell(industry_row, 4).value == "TRUE"
    assert workbook["Flows"].cell(industry_row, 5).value == "TRUE"
    assert workbook["_metadata"]["B4"].value
    import_color_workbook(
        workbook_path,
        config_path,
        custom_path,
        common_rows_path,
        external_colors_path=external_path,
    )
    production = json.loads(config_path.read_text(encoding="utf-8"))
    assert production["product"]["01"] == "#101010"


def test_export_raises_for_missing_external_fuel_colour(tmp_path: Path) -> None:
    config_path = tmp_path / "code_colors.json"
    workbook_path = tmp_path / "colors.xlsx"
    common_rows_path = tmp_path / "common_rows.csv"
    external_path = tmp_path / "colors.json"
    _sample_config(config_path)
    _sample_common_rows(common_rows_path)
    _sample_external_colors(external_path)
    external = json.loads(external_path.read_text(encoding="utf-8"))
    external["fuels"]["standard"] = [
        row for row in external["fuels"]["standard"] if row["code"] != "07.07"
    ]
    external_path.write_text(json.dumps(external), encoding="utf-8")

    with pytest.raises(ValueError, match=r"missing: 07\.07 Diesel"):
        export_color_workbook(
            output_path=workbook_path,
            code_colors_path=config_path,
            common_rows_path=common_rows_path,
            external_colors_path=external_path,
        )
    assert not workbook_path.exists()


def test_synchronize_updates_true_rows_and_preserves_false_rows(tmp_path: Path) -> None:
    config_path = tmp_path / "code_colors.json"
    workbook_path = tmp_path / "colors.xlsx"
    common_rows_path = tmp_path / "common_rows.csv"
    external_path = tmp_path / "colors.json"
    _sample_config(config_path)
    _sample_common_rows(common_rows_path)
    _sample_external_colors(external_path)
    export_color_workbook(
        output_path=workbook_path,
        code_colors_path=config_path,
        common_rows_path=common_rows_path,
        external_colors_path=external_path,
    )
    workbook = load_workbook(workbook_path)
    coal_row = _row_for_key(workbook["Products"], "01")
    electricity_row = _row_for_key(workbook["Products"], "17")
    other_row = _row_for_key(workbook["Other categories"], "LEAP Target")
    workbook["Products"].cell(electricity_row, 4).value = False
    workbook["Products"].cell(electricity_row, 3).value = "#123456"
    workbook["Other categories"].cell(other_row, 3).value = "#654321"
    workbook.save(workbook_path)
    external = json.loads(external_path.read_text(encoding="utf-8"))
    external["fuels"]["standard"][0]["hex"] = "#ABCDEF"
    external_path.write_text(json.dumps(external), encoding="utf-8")
    synchronize_dashboard_colors(
        workbook_path=workbook_path,
        code_colors_path=config_path,
        custom_colors_path=tmp_path / "custom.json",
        common_rows_path=common_rows_path,
        external_colors_path=external_path,
    )
    production = json.loads(config_path.read_text(encoding="utf-8"))
    assert production["product"]["01"] == "#ABCDEF"
    assert production["product"]["17"] == "#123456"
    assert production["plotting"]["series"]["LEAP Target"] == "#654321"
    refreshed = load_workbook(workbook_path)
    assert refreshed["Products"].cell(coal_row, 3).value == "#ABCDEF"
    assert refreshed["Products"].cell(electricity_row, 4).value == "FALSE"


def test_synchronize_true_without_exact_json_uses_components_or_keeps_colour(tmp_path: Path) -> None:
    config_path = tmp_path / "code_colors.json"
    workbook_path = tmp_path / "colors.xlsx"
    common_rows_path = tmp_path / "common_rows.csv"
    external_path = tmp_path / "colors.json"
    _sample_config(config_path)
    _sample_common_rows(common_rows_path)
    _sample_external_colors(external_path)
    axis_components = {
        "product": {},
        "flow": {"01": ("18.01", "18.02", "18.03")},
    }
    export_color_workbook(
        workbook_path,
        config_path,
        common_rows_path,
        external_path,
        axis_components=axis_components,
    )
    workbook = load_workbook(workbook_path)
    unmapped_row = _row_for_key(workbook["Products"], "01.02")
    flow_row = _row_for_key(workbook["Flows"], "01")
    workbook["Products"].cell(unmapped_row, 3).value = "#123456"
    workbook["Products"].cell(unmapped_row, 4).value = True
    workbook.save(workbook_path)

    synchronize_dashboard_colors(
        workbook_path,
        config_path,
        tmp_path / "custom.json",
        common_rows_path,
        external_path,
        axis_components=axis_components,
    )
    production = json.loads(config_path.read_text(encoding="utf-8"))
    assert production["product"]["01.02"] == "#123456"
    assert production["flow"]["01"] == average_oklab(["#362B4F", "#8A5DA7", "#0F266B"])
    refreshed = load_workbook(workbook_path)
    assert refreshed["Products"].cell(unmapped_row, 5).value == "FALSE"
    assert refreshed["Flows"].cell(flow_row, 5).value == "FALSE"
    assert refreshed["Flows"].cell(flow_row, 3).value not in {None, "", "NA", "#N/A"}


def test_metadata_records_used_and_missing_average_components(tmp_path: Path) -> None:
    config_path = tmp_path / "code_colors.json"
    workbook_path = tmp_path / "colors.xlsx"
    common_rows_path = tmp_path / "common_rows.csv"
    _sample_config(config_path)
    _sample_common_rows(common_rows_path)
    axis_components = {
        "product": {},
        "flow": {
            "01": ("18.01", "18.02", "99.99"),
            "14": ("18.01", "18.02"),
        },
    }

    export_color_workbook(
        workbook_path,
        config_path,
        common_rows_path,
        external_colors_path=None,
        sync_flags={"product": {}, "flow": {"01": True, "14": False}},
        axis_components=axis_components,
    )

    workbook = load_workbook(workbook_path)
    metadata = workbook["_metadata"]
    averaged_row = _metadata_row_for_key(metadata, "Flows", "01")
    manual_row = _metadata_row_for_key(metadata, "Flows", "14")
    assert metadata.cell(averaged_row, 7).value == "18.01=#362B4F; 18.02=#8A5DA7"
    assert metadata.cell(averaged_row, 8).value == "99.99"
    assert metadata.cell(manual_row, 7).value is None
    assert metadata.cell(manual_row, 8).value is None


def test_synchronize_is_no_op_when_sources_are_unchanged(tmp_path: Path) -> None:
    config_path = tmp_path / "code_colors.json"
    workbook_path = tmp_path / "dashboard_color_mapping.xlsx"
    common_rows_path = tmp_path / "common_rows.csv"
    external_path = tmp_path / "colors.json"
    custom_path = tmp_path / "custom.json"
    _sample_config(config_path)
    _sample_common_rows(common_rows_path)
    _sample_external_colors(external_path)
    export_color_workbook(workbook_path, config_path, common_rows_path, external_path)
    synchronize_dashboard_colors(workbook_path, config_path, custom_path, common_rows_path, external_path)
    workbook_bytes = workbook_path.read_bytes()

    synchronize_dashboard_colors(workbook_path, config_path, custom_path, common_rows_path, external_path)

    assert workbook_path.read_bytes() == workbook_bytes


def test_synchronize_requires_config_colors_json(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError, match="Place colors.json"):
        synchronize_dashboard_colors(
            workbook_path=tmp_path / "dashboard_color_mapping.xlsx",
            code_colors_path=tmp_path / "code_colors.json",
            custom_colors_path=tmp_path / "custom.json",
            common_rows_path=tmp_path / "common_rows.csv",
            external_colors_path=tmp_path / "missing_colors.json",
        )


def test_self_including_rollup_average_is_stable_across_runs() -> None:
    components = {"product": {}, "flow": {"09.06": ("09.06", "10.01.02", "10.01.03")}}
    component_bases = {"product": {}, "flow": {"09.06": "#3A7CA5"}}
    base = {
        "product": {},
        "flow": {"09.06": "#3A7CA5", "10.01.02": "#3A7CA5", "10.01.03": "#4C8FB5"},
    }
    exact = {"product": {}, "flow": {}}

    first = resolve_json_synced_colors(base, exact, components, component_bases)
    base["flow"]["09.06"] = first["flow"]["09.06"]
    second = resolve_json_synced_colors(base, exact, components, component_bases)

    assert first["flow"]["09.06"] == second["flow"]["09.06"]
    assert first["flow"]["09.06"] == average_oklab(["#3A7CA5", "#3A7CA5", "#4C8FB5"])
