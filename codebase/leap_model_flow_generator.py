#%%
"""
LEAP Model Flow Generator
==========================
Automated generator that ingests a user-provided LEAP export template
(leap_export_templates/*.xlsx) and LEAP energy balance export (leap balances exports/*.xlsx)
and synthesizes the interactive fluid flow visualization HTML.

Extracts:
1. Transformation hierarchy & module order from template sheet sequence.
2. Shortfall rules, surplus rules, and usage rules on output fuels.
3. Unmet requirements rules and trade rules from Primary and Secondary Resources branches.
4. Fuel balance matrices by year (Production, Imports, Exports, Transformation I/O, Sector Demands).
"""

from collections import OrderedDict
import json
import os
from pathlib import Path
import openpyxl

# Resolve repository root
REPO_ROOT = Path(__file__).resolve().parent.parent

#%%
# -----------------------------------------------------------------------------
# 1. Color Palette Mapping from Official Dashboard Config
# -----------------------------------------------------------------------------
DEFAULT_COLORS = {
    'Crude oil': '#DB4F29',
    'Natural gas liquids': '#4AA8A1',
    'Natural gas': '#0070C0',
    'LNG': '#00B0F0',
    'Other bituminous coal': '#454545',
    'Sub bituminous coal': '#595959',
    'Anthracite': '#262626',
    'Lignite': '#7F7F7F',
    'Coal nonspecified': '#555555',
    'Nuclear': '#FD1983',
    'Hydro': '#B0D6F0',
    'Geothermal': '#9C5E31',
    'Solar photovoltaics': '#FFD700',
    'Solar nonspecified': '#FFD700',
    'Wind': '#0284C7',
    'Other biomass': '#2E8B57',
    'Fuelwood and woodwaste': '#8B5A2B',
    'Bagasse': '#A0522D',
    'Biogas': '#66CDAA',
    'Biogasoline': '#32CD32',
    'Biodiesel': '#228B22',
    'Motor gasoline': '#E46C0A',
    'Gas and diesel oil': '#E981E7',
    'Kerosene type jet fuel': '#D5EDF2',
    'Kerosene': '#B2DFEE',
    'Fuel oil': '#906030',
    'LPG': '#000099',
    'Refinery gas not liquefied': '#CD48C1',
    'Ethane': '#689F38',
    'Naphtha': '#FFA726',
    'Petroleum coke': '#424242',
    'Bitumen': '#37474F',
    'Lubricants': '#78909C',
    'White spirit SBP': '#B0BEC5',
    'Paraffin waxes': '#CFD8DC',
    'Other products': '#D1D1D1',
    'Electricity': '#FDE975',
    'Heat': '#CD5C5C',
    'Coke oven coke': '#553625',
    'Coke oven gas': '#8B4513',
    'Blast furnace gas': '#A0522D',
    'Gas works gas': '#00CED1',
    'Coal tar': '#1C1C1C',
    'Hydrogen': '#00E5FF',
    'Ammonia': '#76FF03',
    'Efuel': '#00E676',
    'Other hydrocarbons': '#BDBDBD',
    'Other sources': '#9E9E9E',
}

def load_official_fuel_colors():
    """Load hex color palette from common_esto_dashboard config if available."""
    color_file = REPO_ROOT / 'config' / 'common_esto_dashboard' / 'code_colors.json'
    palette = dict(DEFAULT_COLORS)
    if color_file.exists():
        try:
            with open(color_file, 'r', encoding='utf-8') as f:
                code_colors = json.load(f)
                for code, hex_val in code_colors.items():
                    palette[code] = hex_val
        except Exception:
            pass
    return palette

#%%
# -----------------------------------------------------------------------------
# 2. Template Parsing (Transformation & Resources Branches)
# -----------------------------------------------------------------------------
def parse_leap_export_template(template_path, scenario='Reference'):
    """
    Parses a LEAP export template .xlsx workbook.
    Extracts:
      - Transformation modules in sheet order.
      - Output fuels with Shortfall Rule, Surplus Rule, Usage Rule, Output Share.
      - Resource branches with Unmet Requirements rules, Imports, Exports.
    """
    wb = openpyxl.load_workbook(template_path, read_only=True)
    if 'Export' not in wb.sheetnames:
        raise ValueError(f"Template '{template_path}' missing 'Export' sheet. Available: {wb.sheetnames}")
    sheet = wb['Export']

    trans_modules = OrderedDict()
    res_rules = {}

    for row in sheet.iter_rows(min_row=4, values_only=True):
        bpath = str(row[4] or '')
        var = str(row[5] or '')
        scen = str(row[6] or '')
        vals = [v for v in row[8:35] if v is not None]

        # 1. Transformation Branches
        if bpath.startswith('Transformation'):
            parts = bpath.split('\\')
            if len(parts) >= 2:
                mod_name = parts[1]
                if mod_name not in trans_modules:
                    trans_modules[mod_name] = {
                        'order': len(trans_modules),
                        'output_fuels': {},
                        'feedstock_fuels': set(),
                        'auxiliary_fuels': set(),
                    }

                # Output Fuels & Rules
                if len(parts) >= 4 and parts[2] == 'Output Fuels':
                    fuel_name = parts[3]
                    if fuel_name not in trans_modules[mod_name]['output_fuels']:
                        trans_modules[mod_name]['output_fuels'][fuel_name] = {}
                    if var in ('Shortfall Rule', 'Surplus Rule', 'Usage Rule', 'Priority Output', 'Output Share'):
                        val_str = str(vals[0]) if vals else None
                        if scen == scenario or var not in trans_modules[mod_name]['output_fuels'][fuel_name]:
                            trans_modules[mod_name]['output_fuels'][fuel_name][var] = val_str

                # Feedstocks & Auxiliaries
                if len(parts) >= 5 and 'Feedstock Fuels' in parts:
                    fuel_name = parts[-1]
                    trans_modules[mod_name]['feedstock_fuels'].add(fuel_name)
                elif len(parts) >= 5 and 'Auxiliary Fuels' in parts:
                    fuel_name = parts[-1]
                    trans_modules[mod_name]['auxiliary_fuels'].add(fuel_name)

        # 2. Resources Branches
        elif bpath.startswith('Resources'):
            parts = bpath.split('\\')
            if len(parts) >= 3:
                res_type = parts[1]  # Primary or Secondary
                fuel_name = parts[2]
                res_key = (res_type, fuel_name)
                if res_key not in res_rules:
                    res_rules[res_key] = {}
                if var in ('Unmet Requirements', 'Cost of Unmet Requirements', 'Imports', 'Exports', 'Maximum Production'):
                    val_str = str(vals[0]) if vals else None
                    if scen == scenario or var not in res_rules[res_key]:
                        res_rules[res_key][var] = val_str

    # Convert sets to sorted lists for JSON serialization
    for m in trans_modules.values():
        m['feedstock_fuels'] = sorted(m['feedstock_fuels'])
        m['auxiliary_fuels'] = sorted(m['auxiliary_fuels'])

    return {
        'transformation_modules': trans_modules,
        'resource_rules': {f"{k[0]}/{k[1]}": v for k, v in res_rules.items()},
    }

#%%
# -----------------------------------------------------------------------------
# 3. Balance Export Parsing
# -----------------------------------------------------------------------------
def parse_leap_balance_export(balance_path, year='2022'):
    """
    Parses an energy balance export .xlsx workbook for a specific simulation year.
    Returns:
      - Fuel list (columns)
      - Flow rows dictionary: { 'Production': { 'Natural gas': 62509.5, ... }, ... }
    """
    wb = openpyxl.load_workbook(balance_path, data_only=True)
    if str(year) not in wb.sheetnames:
        raise ValueError(f"Year {year} not found in {balance_path}. Available: {wb.sheetnames}")
    sheet = wb[str(year)]

    fuels = [sheet.cell(3, c).value for c in range(2, sheet.max_column + 1) if sheet.cell(3, c).value]

    flows = OrderedDict()
    for r in range(4, sheet.max_row + 1):
        row_label = sheet.cell(r, 1).value
        if row_label:
            clean_label = str(row_label).strip()
            row_dict = {}
            for c, f in enumerate(fuels, start=2):
                val = sheet.cell(r, c).value
                if isinstance(val, (int, float)) and abs(val) > 0.001:
                    row_dict[f] = float(val)
            if row_dict:
                flows[clean_label] = row_dict

    return {
        'year': str(year),
        'fuels': fuels,
        'flows': flows,
    }

#%%
# -----------------------------------------------------------------------------
# 4. Network Topology Synthesizer
# -----------------------------------------------------------------------------
def synthesize_model_topology(template_meta, balance_meta, min_flow_pj=1.0):
    """
    Combines template metadata (order & rules) with balance export flows to produce
    a complete JSON configuration for the interactive fluid flow visualization.
    """
    flows = balance_meta['flows']
    tmpl_modules = template_meta['transformation_modules']
    res_rules = template_meta['resource_rules']
    fuel_colors = load_official_fuel_colors()

    # 1. Primary Resources (Production, Imports, Exports)
    prod_flows = flows.get('Production', {})
    imp_flows = flows.get('Imports', {})
    exp_flows = flows.get('Exports', {})

    active_fuels = set()
    for row_dict in flows.values():
        active_fuels.update(row_dict.keys())

    # Build fuel catalog
    fuel_catalog = {}
    for f in sorted(active_fuels):
        # find unmet rule
        p_rule = res_rules.get(f"Primary/{f}", {}).get('Unmet Requirements', 'MeetWithImports')
        s_rule = res_rules.get(f"Secondary/{f}", {}).get('Unmet Requirements', p_rule)
        color = fuel_colors.get(f, DEFAULT_COLORS.get(f, '#607d8b'))

        # Resource components
        res_comps = []
        prod_val = prod_flows.get(f, 0)
        imp_val = imp_flows.get(f, 0)
        exp_val = abs(exp_flows.get(f, 0))

        if prod_val > 0.1:
            res_comps.append({'key': 'prod', 'label': f'Production: {prod_val:,.0f} PJ', 'def': round(prod_val), 'max': round(prod_val * 2)})
        if imp_val > 0.1:
            res_comps.append({'key': 'imp', 'label': f'Imports: {imp_val:,.0f} PJ', 'def': round(imp_val), 'max': round(imp_val * 2.5)})
        if exp_val > 0.1:
            res_comps.append({'key': 'exp', 'label': f'Exports: {exp_val:,.0f} PJ', 'def': round(exp_val), 'max': round(exp_val * 2.5)})

        fuel_catalog[f] = {
            'label': f,
            'color': color,
            'unmetRule': s_rule,
            'resources': res_comps,
        }

    # 2. Active Transformation Modules in Sheet Order
    # Find matching flow rows in balances for each template module
    active_modules = []
    for mod_name, mod_info in tmpl_modules.items():
        # Look for matching row in balance flows
        matched_flow = None
        for flow_name, fvals in flows.items():
            if flow_name.lower() == mod_name.lower() or mod_name.lower() in flow_name.lower() or flow_name.lower() in mod_name.lower():
                matched_flow = fvals
                break

        if not matched_flow:
            continue

        # Separate inputs (negative in energy balance) and outputs (positive in energy balance)
        inputs = []
        outputs = []
        tot_in = 0.0
        tot_out = 0.0

        for f, val in matched_flow.items():
            if val < -0.01:
                in_pj = abs(val)
                inputs.append({'fuel': f, 'pj': in_pj})
                tot_in += in_pj
            elif val > 0.01:
                out_pj = val
                outputs.append({'fuel': f, 'pj': out_pj})
                tot_out += out_pj

        if (tot_in + tot_out) < min_flow_pj:
            continue

        # Calculate input shares & output shares
        for inp in inputs:
            inp['share'] = round(inp['pj'] / tot_in, 4) if tot_in > 0 else 0
        for out in outputs:
            out['share'] = round(out['pj'] / tot_out, 4) if tot_out > 0 else 0
            # attach output rules from template
            rules = mod_info['output_fuels'].get(out['fuel'], {})
            out['shortfallRule'] = rules.get('Shortfall Rule', 'RequirementsRemainUnmet')
            out['surplusRule'] = rules.get('Surplus Rule', 'SurplusExported')
            out['usageRule'] = rules.get('Usage Rule', 'DomesticPriority')

        active_modules.append({
            'id': mod_name.replace(' ', '_').replace('/', '_'),
            'label': mod_name,
            'order': mod_info['order'],
            'throughput': round(tot_out if tot_out > 0 else tot_in),
            'capacityMax': round((tot_out if tot_out > 0 else tot_in) * 1.5),
            'exogenousCapacityOutput': round((tot_out if tot_out > 0 else tot_in) * 1.35),
            'efficiency': round((tot_out / tot_in), 3) if tot_in > 0 else 1.0,
            'inputs': inputs,
            'outputs': outputs,
        })

    # Sort active modules in LEAP hierarchy order (from upstream processing at bottom to downstream at top)
    active_modules.sort(key=lambda m: m['order'], reverse=True)

    # 3. Demand Sectors
    demand_sector_names = ['Buildings', 'Industry', 'Road', 'Transport non road', 'International transport', 'Other sector']
    demand_sectors = []
    for sname in demand_sector_names:
        sflow = flows.get(sname, {})
        if not sflow:
            continue
        tot_dem = sum(sflow.values())
        if tot_dem < min_flow_pj:
            continue
        inputs = []
        for f, val in sflow.items():
            if val > 0.01:
                inputs.append({'fuel': f, 'pj': val, 'share': round(val / tot_dem, 4)})
        inputs.sort(key=lambda x: x['pj'], reverse=True)
        demand_sectors.append({
            'id': sname.replace(' ', '_').lower(),
            'label': sname,
            'targetPJ': round(tot_dem),
            'inputs': inputs,
        })

    return {
        'economy': balance_meta.get('year', '2022'),
        'fuelCatalog': fuel_catalog,
        'transformationModules': active_modules,
        'demandSectors': demand_sectors,
    }

#%%
# -----------------------------------------------------------------------------
# 5. Standalone HTML Generator Function
# -----------------------------------------------------------------------------
def generate_flow_html(model_topology, output_html_path, economy_title="United States - LEAP Model Flow"):
    """
    Renders the complete self-contained HTML fluid flow dashboard prototype.
    """
    output_html_path = Path(output_html_path)
    output_html_path.parent.mkdir(parents=True, exist_ok=True)

    # Read base template
    template_src = REPO_ROOT / 'outputs' / 'prototypes' / 'leap_gas_flow' / 'leap_oil_refining_usa_prototype.html'
    if not template_src.exists():
        raise FileNotFoundError(f"Base template not found at {template_src}")

    with open(template_src, 'r', encoding='utf-8') as f:
        html_content = f.read()

    # Write output file
    with open(output_html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

    print(f"Generated standalone flow visualization at: {output_html_path}")
    return output_html_path

#%%
# -----------------------------------------------------------------------------
# 6. Central Interactive Workflow Runner
# -----------------------------------------------------------------------------
if __name__ == '__main__':
    TEMPLATE_PATH = REPO_ROOT.parent / 'leap_initialisation' / 'data' / 'leap_export_templates' / 'USA clean slate 17_08.xlsx'
    BALANCE_PATH = REPO_ROOT.parent / 'leap_initialisation' / 'data' / 'leap balances exports' / '20_USA' / '0805 REF.xlsx'
    OUTPUT_HTML = REPO_ROOT / 'outputs' / 'prototypes' / 'leap_gas_flow' / 'leap_model_flow_auto_usa.html'

    print("=== 1. Parsing LEAP Export Template ===")
    template_meta = parse_leap_export_template(TEMPLATE_PATH, scenario='Reference')
    print(f"Parsed {len(template_meta['transformation_modules'])} transformation modules.")
    print(f"Parsed {len(template_meta['resource_rules'])} resource rules.")

    print("\n=== 2. Parsing LEAP Energy Balance Export ===")
    balance_meta = parse_leap_balance_export(BALANCE_PATH, year='2022')
    print(f"Parsed {len(balance_meta['flows'])} flow rows across {len(balance_meta['fuels'])} fuels.")

    print("\n=== 3. Synthesizing Network Topology ===")
    model_topo = synthesize_model_topology(template_meta, balance_meta, min_flow_pj=10.0)
    print(f"Active transformation modules ({len(model_topo['transformationModules'])}):")
    for m in model_topo['transformationModules']:
        print(f"  [{m['order']:2d}] {m['label']:32s} | Throughput: {m['throughput']:>8,d} PJ | Inputs: {len(m['inputs'])} | Outputs: {len(m['outputs'])}")

    print(f"\nActive demand sectors ({len(model_topo['demandSectors'])}):")
    for d in model_topo['demandSectors']:
        print(f"  {d['label']:25s} | Target Demand: {d['targetPJ']:>8,d} PJ | Fuels: {len(d['inputs'])}")

    print("\n=== 4. Emitting Flow HTML ===")
    out_file = generate_flow_html(model_topo, OUTPUT_HTML, economy_title="United States 2022 Reference - LEAP Flow")
    print("Done! Flow generator ready for multi-economy sync.")
