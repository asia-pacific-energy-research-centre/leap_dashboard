#%%
"""
Expands a single-year LEAP energy balance export into a multi-year file.

Takes a single-year xlsx (any sheet name) and produces a new file with one
sheet per year in the target range. The source year is read from cell A2
(e.g. "Scenario: Target, Year: 2022, Units: Petajoule"). Each output sheet
gets the name "EBal|<year>" and has cell A2 updated to match that year.

Output naming convention: "full model output all years DDMMYYYY <suffix>.xlsx"
"""

import openpyxl
from openpyxl import load_workbook
from copy import copy
import os
import re
from datetime import date
from pathlib import Path


def expand_single_year_ebal(
    input_path: str,
    output_dir: str | None = None,
    year_range: tuple[int, int] = (2023, 2060),
    suffix: str = "REF",
) -> str:
    """
    Expand a single-year EBal sheet into a multi-year workbook.

    Args:
        input_path: Path to the single-year xlsx (e.g. "2022 single year.xlsx").
        output_dir: Directory for the output file. Defaults to same dir as input.
        year_range: (start_year, end_year) inclusive for the sheets to create.
                    The source year sheet is also included automatically.
        suffix: Label appended to the output filename (e.g. "REF", "ALT").

    Returns:
        Path to the created output file.
    """
    src_wb = load_workbook(input_path)
    src_ws = src_wb.active

    # Detect source year from cell A2, e.g. "Scenario: Target, Year: 2022, Units: Petajoule"
    a2_value = src_ws["A2"].value or ""
    match = re.search(r"Year:\s*(\d{4})", a2_value)
    if not match:
        raise ValueError(f"Cannot parse year from cell A2: '{a2_value}'")
    source_year = int(match.group(1))

    # Extract the A2 prefix/suffix so we can reconstruct it for other years
    # e.g. "Scenario: Target, Year: 2022, Units: Petajoule"
    a2_template = re.sub(r"(Year:\s*)\d{4}", r"\g<1>{year}", a2_value)

    start_year, end_year = year_range
    all_years = sorted(
        set(range(start_year, end_year + 1)) | {source_year}, reverse=True
    )

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)  # remove default empty sheet

    # Cache source data once
    src_data = list(src_ws.iter_rows(values_only=True))
    src_col_dims = {
        col: src_ws.column_dimensions[col].width
        for col in src_ws.column_dimensions
    }
    src_row_dims = {
        row: src_ws.row_dimensions[row].height
        for row in src_ws.row_dimensions
    }

    for year in all_years:
        sheet_name = f"EBal|{year}"
        ws = out_wb.create_sheet(title=sheet_name)

        # Write values, updating A2 with the correct year
        for r_idx, row in enumerate(src_data, start=1):
            for c_idx, value in enumerate(row, start=1):
                if r_idx == 2 and c_idx == 1:
                    value = a2_template.format(year=year)
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Preserve column widths
        for col, width in src_col_dims.items():
            ws.column_dimensions[col].width = width

        # Preserve row heights
        for row_num, height in src_row_dims.items():
            ws.row_dimensions[row_num].height = height

    today = date.today().strftime("%d%m%Y")
    filename = f"full model output all years {today} {suffix}.xlsx"
    if output_dir is None:
        output_dir = os.path.dirname(input_path)
    output_path = os.path.join(output_dir, filename)

    out_wb.save(output_path)
    src_wb.close()
    print(f"Created {len(all_years)} sheets ({all_years[-1]}–{all_years[0]}) -> {output_path}")
    return output_path


if __name__ == "__main__":
    canonical_root = Path(
        os.environ.get(
            "LEAP_BALANCE_EXPORTS_ROOT",
            Path(__file__).resolve().parents[3]
            / "leap_initialisation"
            / "data"
            / "leap balances exports",
        )
    )
    INPUT = str(canonical_root / "20_USA" / "2022 28052026 tgt single year.xlsx")

    expand_single_year_ebal(
        input_path=INPUT,
        year_range=(2023, 2060),
        suffix="TGT",
    )
#%%
