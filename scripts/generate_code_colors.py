"""Generate config/common_esto_dashboard/code_colors.json.

This script is the source of truth for the dashboard's fuel and flow colours;
edit it and re-run rather than hand-editing the generated JSON, which carries
no record of why any given colour was chosen.

Anchors marked "(legend)" are carried over verbatim from the previous
dashboard's product_color_legend (config/archive/leap_comparison_dashboard_
template_v4.json), which keyed colours by display name. They are re-keyed to
ESTO codes here: a common ESTO label takes its name from the first component of
its partition, so the name moves when a rollup changes while the code span
stays put. Siblings the old legend never named are shaded around their family
anchor so a stacked chart can still tell them apart.

Run from the repo root:  python scripts/generate_code_colors.py
"""
from __future__ import annotations

import colorsys
import json
import os
import re
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from codebase.dashboard_color_config import build_common_rollup_colors, load_common_rollup_memberships

OUT = REPO_ROOT / "config" / "common_esto_dashboard" / "code_colors.json"
CUSTOM_COLORS = OUT.with_name("code_colors_custom.json")
COLOR_SOURCE_XLSX = OUT.parents[1] / "archive" / "master_config 9th visualisation.xlsx"
DEFAULT_COMMON_ROWS = REPO_ROOT.parent / "leap_mappings" / "results" / "common_esto" / "common_esto_rows.csv"
FALLBACK_COMMON_ROWS = REPO_ROOT / "tests" / "fixtures" / "common_esto_dashboard" / "common_esto_rows.csv"
COMMON_ROWS = Path(os.getenv("COMMON_ESTO_ROWS_PATH", str(DEFAULT_COMMON_ROWS)))
if not COMMON_ROWS.exists():
    COMMON_ROWS = FALLBACK_COMMON_ROWS
try:
    COMMON_ROWS_SOURCE_LABEL = COMMON_ROWS.relative_to(REPO_ROOT.parent).as_posix()
except ValueError:
    COMMON_ROWS_SOURCE_LABEL = str(COMMON_ROWS)


def _hls(base: str) -> tuple[float, float, float]:
    r, g, b = (int(base[i:i + 2], 16) / 255 for i in (1, 3, 5))
    return colorsys.rgb_to_hls(r, g, b)


def _hex(h: float, l: float, s: float) -> str:
    r, g, b = colorsys.hls_to_rgb(h % 1.0, l, s)
    return "#{:02X}{:02X}{:02X}".format(round(r * 255), round(g * 255), round(b * 255))


# Lightness window that stays readable as a filled area on a white background:
# above this siblings wash out, below it they read as black.
_L_MIN, _L_MAX = 0.24, 0.78


def family(base: str, codes: list[str], hue_drift: float = 0.05) -> dict[str, str]:
    """Spread siblings over a lightness ramp plus a slight hue drift.

    An earlier version offset lightness by a fixed amount and clamped at the
    ends, which silently collapsed large families (all of 07.01-07.03 came out
    the same colour). Spacing the ramp evenly inside the readable window
    guarantees every sibling differs, and the hue drift keeps long families
    (the 17 petroleum products) separable rather than a single-hue gradient.
    """
    if len(codes) == 1:
        return {codes[0]: base}
    h, _, s = _hls(base)
    n = len(codes)
    return {
        code: _hex(h + hue_drift * (2 * i / (n - 1) - 1), _L_MIN + (_L_MAX - _L_MIN) * i / (n - 1), s)
        for i, code in enumerate(codes)
    }


def _read_color_source() -> tuple[dict[str, str], list[str]]:
    """Read the archived plotting-name colour catalogue.

    Duplicate names with different colours are intentionally excluded from
    fallback resolution because the workbook does not provide enough context
    to choose safely between them.
    """
    workbook = openpyxl.load_workbook(COLOR_SOURCE_XLSX, read_only=True, data_only=True)
    rows = workbook["colors"].iter_rows(values_only=True)
    next(rows, None)  # header
    colors_by_name: dict[str, set[str]] = {}
    for row in rows:
        if not row or row[0] is None or row[1] is None:
            continue
        name = str(row[0]).strip()
        color = str(row[1]).strip()
        if name and color.startswith("#"):
            colors_by_name.setdefault(name, set()).add(color)
    ambiguous = sorted(name for name, colors in colors_by_name.items() if len(colors) > 1)
    colors = {name: next(iter(values)) for name, values in colors_by_name.items() if len(values) == 1}
    return colors, ambiguous


def _sheet_values(sheet_name: str, column_name: str) -> list[str]:
    workbook = openpyxl.load_workbook(COLOR_SOURCE_XLSX, read_only=True, data_only=True)
    rows = workbook[sheet_name].iter_rows(values_only=True)
    header = [str(value).strip() if value is not None else "" for value in next(rows)]
    column_index = header.index(column_name)
    return sorted({str(row[column_index]).strip() for row in rows if row[column_index] is not None and str(row[column_index]).strip()})


def _plotting_color_catalogue(source_colors: dict[str, str]) -> tuple[dict[str, dict[str, str]], dict[str, list[str]]]:
    """Cross-check plotting categories and retain source colours by axis.

    The workbook has one shared colour sheet, but its plotting mappings define
    separate fuel, sector/transformation, and capacity namespaces.
    """
    names_by_axis = {
        "product": _sheet_values("fuels_plotting", "fuels_plotting"),
        "flow": sorted(set(
            _sheet_values("sectors_plotting", "sectors_plotting")
            + _sheet_values("transformation_sector_mappings", "sectors_plotting")
        )),
        "capacity": _sheet_values("capacity_plotting", "capacity_plotting"),
    }
    source_by_folded_name = {name.casefold(): color for name, color in source_colors.items()}
    catalogue: dict[str, dict[str, str]] = {}
    missing: dict[str, list[str]] = {}
    for axis, names in names_by_axis.items():
        catalogue[axis] = {}
        missing[axis] = []
        for name in names:
            color = source_colors.get(name) or source_by_folded_name.get(name.casefold())
            if color:
                catalogue[axis][name] = color
            else:
                missing[axis].append(name)
    return catalogue, missing


def _numeric_code(value: object) -> str:
    """Convert a workbook hierarchy label such as ``09_01_power`` to 09.01."""
    text = str(value or "").strip()
    parts = re.match(r"^(\d+(?:_\d+)*)_", text)
    if not parts:
        return ""
    return ".".join(parts.group(1).split("_"))


def _add_mapping_source_codes(colors: dict[str, str], sheet_name: str, category_column: str, source_colors: dict[str, str]) -> None:
    """Use mapped plotting colours only to fill code entries still absent."""
    workbook = openpyxl.load_workbook(COLOR_SOURCE_XLSX, read_only=True, data_only=True)
    rows = workbook[sheet_name].iter_rows(values_only=True)
    header = [str(value).strip() if value is not None else "" for value in next(rows)]
    category_index = header.index(category_column)
    code_columns = [index for index, value in enumerate(header) if value in {"sectors", "sub1sectors", "sub2sectors", "sub3sectors", "sub4sectors", "fuels", "subfuels"}]
    source_by_folded_name = {name.casefold(): color for name, color in source_colors.items()}
    for row in rows:
        category = str(row[category_index] or "").strip()
        color = source_colors.get(category) or source_by_folded_name.get(category.casefold())
        if not color:
            continue
        for index in code_columns:
            code = _numeric_code(row[index])
            if code:
                colors.setdefault(code, color)


# --- products -------------------------------------------------------------
# Anchors marked (legend) are verbatim from leap_comparison_dashboard_template_v4.
product: dict[str, str] = {}
product["01"] = "#0D0D0D"           # Coal (legend)
product["01.01"] = "#454545"        # Coking coal == legend "Metallurgical coal"
product["01.02"] = "#A6A6A6"        # Other bituminous == legend "Thermal coal"
product["01.03"] = "#C4C4C4"        # Sub-bituminous
product["01.04"] = "#7D7D7D"        # Anthracite
product["01.05"] = "#8C0000"        # Lignite (legend)
product["01.99"] = "#595959"
product["02"] = "#C1A470"           # Coal products (legend)
product.update(family("#C1A470", ["02.01", "02.02", "02.03", "02.04",
                                  "02.05", "02.06", "02.07", "02.08"]))
product["03"] = "#7B5E3B"           # Peat - no legend precedent
product["04"] = "#9C8055"           # Peat products - no legend precedent
product["06"] = "#2E6864"           # legend "Crude, NGLs & other hydrocarbons"
product["06.01"] = "#DB4F29"        # Crude oil (legend)
product["06.02"] = "#4AA8A1"        # Natural gas liquids (legend)
product["06.03"] = "#2E6864"
product["06.04"] = "#6FC2BB"
product["06.05"] = "#8FD4CE"
product["07"] = "#842482"           # Petroleum products (legend)
product.update(family("#842482", [f"07.{i:02d}" for i in range(1, 18)], hue_drift=0.09))
product["07.99"] = "#842482"
product["08"] = "#0070C0"           # Natural gas (legend)
product["08.01"] = "#0070C0"
product["08.02"] = "#A20042"        # LNG (legend)
product["08.03"] = "#4DA3DB"
product["08.99"] = "#6BB5E4"
product["09"] = "#C6188C"           # Nuclear (legend)
product["10"] = "#B0D6F0"           # Hydro (legend)
product["11"] = "#9C5E31"           # Geothermal (legend)
product["12"] = "#FFD700"           # Solar (legend)
product["12.01"] = "#FFD700"
product["12.99"] = "#FFE84D"
product["13"] = "#ABD7F5"           # Ocean (legend)
product["14"] = "#000099"           # Wind (legend)
product["15"] = "#2E8B57"           # Solid biomass (legend)
product["15.01"] = "#2E8B57"        # Fuelwood == legend "Solid biomass"
product["15.02"] = "#4CA373"        # Bagasse
product["15.03"] = "#1F5F3C"        # Charcoal
product["15.04"] = "#6BBA8F"        # Black liquor
product["15.05"] = "#8ACFAA"        # Other biomass
product["16"] = "#8A8A8A"           # Others (legend)
product["16.01"] = "#00C25B"        # Biogas; legend #00FE73 darkened for white bg
product["16.02"] = "#7F6A55"        # Industrial waste
product["16.03"] = "#A88B68"        # MSW renewable
product["16.04"] = "#5C5148"        # MSW non-renewable
product["16.05"] = "#F09417"        # Biogasoline (legend)
product["16.06"] = "#304A1E"        # Biodiesel (legend)
product["16.07"] = "#7FA828"        # Bio jet kerosene; legend #9ACD32 darkened
product["16.08"] = "#8FAF3D"        # Other liquid biofuels
product["16.09"] = "#8A8A8A"        # Other sources
product["16.10"] = "#5FB3A3"        # Ammonia - no legend precedent
product["16.11"] = "#D46FA0"        # legend "Hydrogen-based fuels"
product["16.12"] = "#F67AA3"        # Hydrogen (legend)
product["17"] = "#FFD757"           # Electricity (legend)
product["18"] = "#DC143C"           # Heat (legend)

# --- flows ----------------------------------------------------------------
# No archived precedent. Sector anchors reuse the renderer's existing
# sector_colors hues so the overview and the sector pages agree.
flow: dict[str, str] = {}
flow["01"] = "#2F855A"              # Production
flow["02"] = "#3B6FB6"              # Imports
flow["03"] = "#B64C67"              # Exports
flow["04"] = "#007C78"              # International marine bunkers
flow["05"] = "#5EAAD4"              # International aviation bunkers
flow["06"] = "#9A6B2F"              # Stock changes
flow["07"] = "#486581"              # TPES
flow["08"] = "#7B5E57"              # Transfers
flow["09"] = "#7A5195"              # Total transformation
flow["09.01"] = "#6B4586"           # Main activity producer
flow["09.02"] = "#4F46E5"           # Autoproducer - own hue, see below
flow.update(family("#7A5195", ["09.01.01", "09.01.02", "09.01.03"]))
# Autoproducer plants carry the same display names as their main-activity
# counterparts (09.02.01 and 09.01.01 are both "Electricity plants") and
# stack in the same chart, so they get a separate hue rather than a shared
# ramp that would resolve both to one colour.
flow.update(family("#4F46E5", ["09.02.01", "09.02.02", "09.02.03"]))
flow["09.03"] = "#B99BCC"           # Heat pumps
flow["09.04"] = "#9D7BB5"           # Electric boilers
flow["09.05"] = "#C7B0D6"           # Chemical heat
flow["09.06"] = "#3A7CA5"           # Gas processing
flow.update(family("#3A7CA5", ["09.06.01", "09.06.02", "09.06.03", "09.06.04"]))
flow["09.06.02.01"] = "#2C6180"     # Liquefaction
flow["09.06.02.02"] = "#5FA3C9"     # Regasification
flow["09.07"] = "#C65D28"           # Oil refineries
flow["09.08"] = "#6E6E6E"           # Coal transformation
flow.update(family("#6E6E6E", ["09.08.01", "09.08.02", "09.08.03", "09.08.04", "09.08.05"]))
flow["09.09"] = "#A63D40"           # Petrochemical industry
flow["09.11"] = "#5A7D3A"           # Charcoal processing
flow["09.12"] = "#8C8C8C"           # Non-specified transformation
flow["09.13"] = "#F67AA3"           # Hydrogen transformation
flow["09.13.01"] = "#E2537F"
flow["09.13.03"] = "#F79CBB"
flow["10"] = "#9A6B2F"              # Own use and losses
flow["10.01"] = "#9A6B2F"
flow["10.01.01"] = "#7A5195"        # Electricity, CHP and heat plants
flow["10.01.02"] = "#3A7CA5"        # Gas works plants
flow["10.01.03"] = "#4C8FB5"        # Liquefaction/regasification
flow["10.01.05"] = "#6E6E6E"        # Coke ovens
flow["10.01.06"] = "#0D0D0D"        # Coal mines
flow["10.01.07"] = "#545454"        # Blast furnaces
flow["10.01.11"] = "#C65D28"        # Oil refineries
flow["10.01.12"] = "#DB4F29"        # Oil and gas extraction
flow["10.01.13"] = "#B0D6F0"        # Pump storage plants
flow["10.01.17"] = "#8C8C8C"        # Non-specified own uses
flow["10.02"] = "#D08496"           # Transmission and distribution losses - clear of 03 Exports
flow["11"] = "#9CA3AF"              # Statistical discrepancy - clear of 09.08 coal transformation
flow["12"] = "#6B7FA3"              # TFC - kept clear of 07 TPES, which it stacks beside
flow["13"] = "#8494B3"              # TFEC
flow["14"] = "#3B82F6"              # Industry (renderer sector_colors)
flow["14.01"] = "#1E40AF"           # Mining and quarrying
flow["14.02"] = "#0EA5E9"           # Construction - clear of the 14.03.xx ramp
flow["14.03"] = "#3B82F6"           # Manufacturing
flow.update(family("#3B82F6", [f"14.03.{i:02d}" for i in range(1, 12)], hue_drift=0.08))
flow["15"] = "#F97316"              # Transport (renderer sector_colors)
flow.update(family("#F97316", ["15.01", "15.02", "15.03", "15.04", "15.05", "15.06"]))
flow["16.01"] = "#10B981"           # Commercial - buildings (renderer sector_colors)
flow["16.01.01"] = "#0B8F63"        # Datacentres
flow["16.01.99"] = "#4FD1A5"        # Commercial unallocated
flow["16.02"] = "#0E9F73"           # Residential - buildings
flow["16.03"] = "#8B5CF6"           # Agriculture - others (renderer sector_colors)
flow["16.04"] = "#A78BFA"           # Fishing - others
flow["16.05"] = "#6D3FD4"           # Non-specified others
flow["17"] = "#94A3B8"              # Non-energy use (renderer sector_colors)
# The 9th's 18.x/19.x plant codes each mirror an ESTO 09.01.x (main activity)
# or 09.02.x (autoproducer) plant, so they take their twin's colour outright:
# the same plant should look the same whichever source names it.
flow["18"] = "#7A5195"              # 9th MAP/AP electricity plants
flow["18.01"] = flow["09.01.01"]    # MAP electricity  ~ MAP electricity plants
flow["18.02"] = flow["09.01.02"]    # MAP CHP          ~ MAP CHP plants
flow["18.03"] = flow["09.02.01"]    # AP electricity   ~ AP electricity plants
flow["18.04"] = flow["09.02.02"]    # AP CHP           ~ AP CHP plants
flow["19"] = "#C65D28"              # 9th MAP/AP heat plants
flow["19.01"] = flow["09.01.02"]    # MAP CHP          ~ MAP CHP plants
flow["19.02"] = flow["09.01.03"]    # MAP heat         ~ MAP heat plants
flow["19.03"] = flow["09.02.02"]    # AP CHP           ~ AP CHP plants
flow["19.04"] = flow["09.02.03"]    # AP heat          ~ AP heat plants

# Fill only gaps in the maintained code map from the archived plotting
# catalogue. Existing code choices remain authoritative.
source_colors, ambiguous_source_colors = _read_color_source()
plotting_colors, plotting_colors_missing = _plotting_color_catalogue(source_colors)
_add_mapping_source_codes(product, "fuels_plotting", "fuels_plotting", source_colors)
_add_mapping_source_codes(flow, "sectors_plotting", "sectors_plotting", source_colors)
_add_mapping_source_codes(flow, "transformation_sector_mappings", "sectors_plotting", source_colors)

# A colleague-edited Excel workbook can supply a complete custom layer through
# scripts/manage_dashboard_colors.py. Apply it last so regenerating the base
# catalogue never discards reviewed choices.
custom_colors = json.loads(CUSTOM_COLORS.read_text(encoding="utf-8")) if CUSTOM_COLORS.exists() else {}
product.update(dict(custom_colors.get("product", {})))
flow.update(dict(custom_colors.get("flow", {})))
for axis, colors in dict(custom_colors.get("plotting", {})).items():
    plotting_colors.setdefault(axis, {}).update(dict(colors))
common_memberships = load_common_rollup_memberships(COMMON_ROWS)
common_colors = build_common_rollup_colors(
    {"product": product, "flow": flow},
    common_memberships,
    overrides=dict(custom_colors.get("common_overrides", {})),
)

payload = {
    "_generated_by": "scripts/generate_code_colors.py - edit that script, not this file",
    "_color_source": "config/archive/master_config 9th visualisation.xlsx, colors sheet",
    "_custom_color_source": CUSTOM_COLORS.name if CUSTOM_COLORS.exists() else "",
    "_common_color_source": COMMON_ROWS_SOURCE_LABEL,
    "_common_color_method": "equal-weight OKLab average of mapping-owned ESTO components",
    "_common_color_overrides": dict(custom_colors.get("common_overrides", {})),
    "_ambiguous_source_colors": ambiguous_source_colors,
    "_source_plotting_colors": dict(sorted(source_colors.items())),
    "_plotting_color_coverage": {
        axis: {
            "mapped": len(plotting_colors[axis]),
            "missing": plotting_colors_missing[axis],
        }
        for axis in plotting_colors
    },
    "_notes": [
        "Maps ESTO code -> hex, per axis. Keyed by code, never by display name:",
        "Exact Common ESTO categories use their configured ESTO code colour.",
        "Multi-component Common ESTO categories use an equal-weight OKLab",
        "average of their mapping-owned component colours. Missing exact",
        "component colours inherit from the nearest configured code parent.",
        "Product and flow codes are separate namespaces and must not be merged:",
        "product 16 is Others, flow 16.01 is Commercial and public services.",
        "Anchors marked (legend) in the generator match the previous dashboard's",
        "product_color_legend so charts stay recognisable across the rebuild.",
        "Some codes intentionally share a colour: a rolled row (01.02-01.04) with",
        "its unrolled component, and the same facility seen from two code spaces",
        "(09.07 Oil refineries / 10.01.11 Oil refineries, ESTO 09.01.01 / 9th 18.01).",
    ],
    "plotting": plotting_colors,
    "product": dict(sorted(product.items())),
    "flow": dict(sorted(flow.items())),
    "common": {
        axis: dict(sorted(colors.items()))
        for axis, colors in common_colors.items()
    },
}
OUT.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
print(f"wrote {OUT} - {len(product)} product codes, {len(flow)} flow codes")
