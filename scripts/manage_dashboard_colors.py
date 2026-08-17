#%%
"""Export and import the non-technical Excel editor for dashboard colours.

The workbook is a human editing surface. Production continues to read
``code_colors.json``; importing a returned workbook writes a complete custom
colour layer and applies it to that production file.
"""
from __future__ import annotations

import hashlib
import json
import re
import sys
from colorsys import hls_to_rgb, rgb_to_hls
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from codebase.dashboard_color_config import (
    average_oklab,
    build_common_rollup_colors,
    load_common_rollup_memberships,
    normalize_hex,
)
from codebase.hierarchy_subtotal_contract_loader import load_hierarchy_subtotal_contract


# --- Stable paths and workbook contract -----------------------------------

CONFIG_DIR = REPO_ROOT / "config" / "common_esto_dashboard"
CODE_COLORS_PATH = CONFIG_DIR / "code_colors.json"
CUSTOM_COLORS_PATH = CONFIG_DIR / "code_colors_custom.json"
DEFAULT_WORKBOOK_PATH = REPO_ROOT / "outputs" / "dashboard_color_mapping" / "dashboard_color_mapping.xlsx"
DEFAULT_EXTERNAL_COLORS_PATH = CONFIG_DIR / "colors.json"
HIERARCHY_CONTRACT_DIR = (
    REPO_ROOT.parent / "leap_mappings" / "results" / "hierarchy_subtotal_contract" / "current"
)
MAPPINGS_WORKBOOK_PATH = REPO_ROOT.parent / "leap_mappings" / "config" / "outlook_mappings_master.xlsx"
UPSTREAM_AXIS_NODES_PATH = (
    REPO_ROOT.parent
    / "leap_mappings"
    / "results"
    / "hierarchy_subtotal_contract"
    / "current"
    / "axis_nodes.csv"
)
FALLBACK_COMMON_ROWS_PATH = REPO_ROOT / "tests" / "fixtures" / "common_esto_dashboard" / "common_esto_rows.csv"
UPSTREAM_COMMON_ROWS_PATH = REPO_ROOT.parent / "leap_mappings" / "results" / "common_esto" / "common_esto_rows.csv"
DEFAULT_COMMON_ROWS_PATH = UPSTREAM_COMMON_ROWS_PATH if UPSTREAM_COMMON_ROWS_PATH.exists() else FALLBACK_COMMON_ROWS_PATH

WORKBOOK_SCHEMA_VERSION = "6"
PLACEHOLDER_LABEL = "Category not in the current common hierarchy"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
LOCKED_FILL = PatternFill("solid", fgColor="F2F2F2")
WHITE_FONT = Font(color="FFFFFF")
BLACK_FONT = Font(color="000000")

SHEET_SPECS = (
    ("Products", "product"),
    ("Flows", "flow"),
)
OTHER_SHEET_NAME = "Other categories"

# The APERC colour tool uses an EBT plotting-code space that is not identical
# to Common ESTO. Keep the semantic bridge explicit: never assume that equal
# numeric codes mean equal fuels (notably the 16.x families).
REQUIRED_EXTERNAL_FUELS = {
    "01": "Coal",
    "06": "Oil",
    "07.01": "Gasoline",
    "07.02": "AvGas",
    "07.03": "Naphtha",
    "07.05": "KeroJet",
    "07.06": "OthKer",
    "07.07": "Diesel",
    "07.08": "FuelOil",
    "07.09": "LPG",
    "07.11": "Ethane",
    "07.14": "Bitumen",
    "07.17": "OthProd",
    "07.97": "Other Oil",
    "08": "Gas",
    "09": "Nuclear",
    "10": "Hydro",
    "11": "Geothermal",
    "12": "Solar",
    "14": "Wind",
    "15": "Biomass",
    "16.01": "Other RE",
    "16.02": "Others",
    "17": "Electricity",
    "17.20": "Electricity (Imported)",
    "18": "Heat",
}
EXTERNAL_FUEL_TO_PRODUCT_CODES = {
    "01": ("01", "02"),
    "06": ("06", "07"),
    "07.01": ("07.01",),
    "07.02": ("07.02",),
    "07.03": ("07.03",),
    "07.05": ("07.05",),
    "07.06": ("07.06",),
    "07.07": ("07.07",),
    "07.08": ("07.08",),
    "07.09": ("07.09",),
    "07.11": ("07.11",),
    "07.14": ("07.14",),
    "07.17": ("07.17",),
    "07.97": ("07.99",),
    "08": ("08",),
    "09": ("09",),
    "10": ("10",),
    "11": ("11",),
    "12": ("12",),
    "14": ("14",),
    "15": ("15",),
    "16.01": ("13",),
    "16.02": ("16",),
    "17": ("17",),
    "17.20": (),
    "18": ("18",),
}
REQUIRED_EXTERNAL_SECTORS = {
    "14": "Industry",
    "15": "Transport",
    "16.01": "Commercial",
    "16.02": "Residential",
    "16.03": "Agriculture & Others",
    "17": "Non-energy",
}
EXTERNAL_SECTOR_TO_FLOW_CODES = {
    code: (code,) for code in REQUIRED_EXTERNAL_SECTORS
}


# --- Colour and label helpers ---------------------------------------------

def _cell_fill_hex(cell: object) -> str:
    """Read an ordinary solid Excel fill as #RRGGBB, or return blank."""
    fill = getattr(cell, "fill", None)
    foreground = getattr(fill, "fgColor", None)
    if getattr(fill, "fill_type", None) != "solid" or foreground is None:
        return ""
    color_type = getattr(foreground, "type", "")
    if color_type == "rgb" and isinstance(foreground.rgb, str):
        rgb = foreground.rgb[-6:]
        return f"#{rgb.upper()}" if re.fullmatch(r"[0-9A-Fa-f]{6}", rgb) else ""
    if color_type == "indexed" and isinstance(foreground.indexed, int):
        if 0 <= foreground.indexed < len(COLOR_INDEX):
            return f"#{COLOR_INDEX[foreground.indexed][-6:].upper()}"
    if color_type == "theme" and isinstance(foreground.theme, int):
        return _theme_fill_hex(cell, foreground.theme, float(foreground.tint or 0))
    return ""


def _theme_fill_hex(cell: object, theme_index: int, tint: float) -> str:
    """Resolve an Excel theme fill, including its light/dark tint."""
    workbook = getattr(getattr(cell, "parent", None), "parent", None)
    theme_xml = getattr(workbook, "loaded_theme", None)
    if not theme_xml:
        return ""
    root = ElementTree.fromstring(theme_xml)
    namespace = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
    scheme = root.find(f".//{namespace}clrScheme")
    if scheme is None or not 0 <= theme_index < len(scheme):
        return ""
    color_node = next(iter(scheme[theme_index]), None)
    if color_node is None:
        return ""
    rgb = color_node.attrib.get("val") or color_node.attrib.get("lastClr") or ""
    if not re.fullmatch(r"[0-9A-Fa-f]{6}", rgb):
        return ""
    red, green, blue = (int(rgb[index:index + 2], 16) / 255 for index in (0, 2, 4))
    hue, lightness, saturation = rgb_to_hls(red, green, blue)
    lightness = lightness * (1 + tint) if tint < 0 else lightness * (1 - tint) + tint
    tinted = hls_to_rgb(hue, max(0, min(1, lightness)), saturation)
    return "#{:02X}{:02X}{:02X}".format(*(round(channel * 255) for channel in tinted))


def _font_for_fill(hex_color: str) -> Font:
    """Use white text on dark fills and black text on light fills."""
    color = normalize_hex(hex_color)
    red, green, blue = (int(color[index:index + 2], 16) for index in (1, 3, 5))
    luminance = (0.2126 * red + 0.7152 * green + 0.0722 * blue) / 255
    return WHITE_FONT if luminance < 0.46 else BLACK_FONT


def _load_axis_labels() -> dict[str, dict[str, str]]:
    """Load display labels from mapping-owned outputs without deriving hierarchy."""
    import csv

    labels: dict[str, dict[str, str]] = {"product": {}, "flow": {}}
    if UPSTREAM_AXIS_NODES_PATH.exists():
        with UPSTREAM_AXIS_NODES_PATH.open(encoding="utf-8-sig", newline="") as handle:
            axis_rows = list(csv.DictReader(handle))
        # Native ESTO/extended nodes supply friendly names for configured
        # parent or extension codes absent from the current common hierarchy.
        # Common ESTO labels then overwrite them where the contract has one.
        for dataset_id in ("esto_extended", "esto", "common_esto"):
            for row in axis_rows:
                if row.get("dataset_id") != dataset_id or row.get("axis_role") not in labels:
                    continue
                node_label = str(row.get("node_label", "")).strip()
                code, _, name = node_label.partition(" ")
                if code:
                    labels[str(row["axis_role"])][code] = name or node_label
        return labels

    if FALLBACK_COMMON_ROWS_PATH.exists():
        with FALLBACK_COMMON_ROWS_PATH.open(encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                for axis in ("product", "flow"):
                    code = str(row.get(f"component_{axis}_code", "")).strip()
                    name = str(row.get(f"component_{axis}_name", "")).strip()
                    if code and name:
                        labels[axis].setdefault(code, name)
    return labels


def _config_hash(payload: dict[str, object]) -> str:
    serialized = json.dumps(payload, sort_keys=True, ensure_ascii=False).encode("utf-8")
    return hashlib.sha256(serialized).hexdigest()


def _file_hash(path: Path) -> str:
    """Return a stable hash for an external source file."""
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _load_external_rows(
    payload: dict[str, object],
    section: str,
    collection: str,
    required: dict[str, str],
) -> dict[str, str]:
    """Validate one code/hex collection in the APERC colour JSON."""
    rows = dict(payload.get(section, {})).get(collection)
    if not isinstance(rows, list):
        raise ValueError(f"APERC colors.json must contain a {section}.{collection} list")
    colors: dict[str, str] = {}
    invalid_rows: list[str] = []
    for row_number, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            invalid_rows.append(f"row {row_number}: expected an object")
            continue
        code = str(row.get("code", "")).strip()
        label = str(row.get("label", "")).strip()
        if not code:
            invalid_rows.append(f"row {row_number}: missing code")
            continue
        if code in colors:
            invalid_rows.append(f"row {row_number}: duplicate code {code}")
            continue
        try:
            colors[code] = normalize_hex(row.get("hex"))
        except ValueError:
            invalid_rows.append(f"{code} {label or '(unnamed)'}: missing or invalid hex colour")
    missing_required = [f"{code} {label}" for code, label in required.items() if code not in colors]
    if invalid_rows or missing_required:
        details = [
            *(f"invalid: {message}" for message in invalid_rows),
            *(f"missing: {message}" for message in missing_required),
        ]
        raise ValueError(
            f"APERC colors.json does not provide every required {section} colour:\n- "
            + "\n- ".join(details)
        )
    return colors


def load_external_sync_colors(external_colors_path: Path) -> tuple[dict[str, dict[str, str]], str]:
    """Read the authoritative detailed fuel and end-use sector palettes."""
    if not external_colors_path.exists():
        raise FileNotFoundError(
            f"APERC colour file not found: {external_colors_path}. "
            "Place colors.json in config/common_esto_dashboard before running the dashboard."
        )
    try:
        payload = json.loads(external_colors_path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"Could not read APERC colour file {external_colors_path}: {exc}") from exc

    fuels = _load_external_rows(payload, "fuels", "standard", REQUIRED_EXTERNAL_FUELS)
    sectors = _load_external_rows(payload, "sectors", "list", REQUIRED_EXTERNAL_SECTORS)
    sync_colors: dict[str, dict[str, str]] = {"product": {}, "flow": {}}
    for external_code, color in fuels.items():
        targets = EXTERNAL_FUEL_TO_PRODUCT_CODES.get(external_code)
        if targets is None and external_code.startswith("07."):
            targets = (external_code,)
        for target in targets or ():
            sync_colors["product"][target] = color
    for external_code, color in sectors.items():
        for target in EXTERNAL_SECTOR_TO_FLOW_CODES.get(external_code, ()):
            sync_colors["flow"][target] = color
    return sync_colors, _file_hash(external_colors_path)


def _code_from_axis_label(label: object) -> str:
    """Return the stable leading code from a mapping-owned axis label."""
    return str(label or "").strip().partition(" ")[0]


def load_axis_color_components(
    hierarchy_contract_dir: Path = HIERARCHY_CONTRACT_DIR,
    mappings_workbook_path: Path = MAPPINGS_WORKBOOK_PATH,
) -> dict[str, dict[str, tuple[str, ...]]]:
    """Load declared hierarchy and rollup members used for colour averaging.

    Parenthood comes from the published hierarchy contract. Standalone
    comparison-boundary rollups (including own use) come from the upstream
    mapping implementation and its authoritative rollup sheet.
    """
    components: dict[str, dict[str, list[str]]] = {"product": {}, "flow": {}}
    if hierarchy_contract_dir.exists():
        _manifest, frames = load_hierarchy_subtotal_contract(hierarchy_contract_dir)
        axis_nodes = frames["axis_nodes"]
        common_nodes = axis_nodes[axis_nodes["dataset_id"].astype(str) == "common_esto"]
        labels_by_id = {
            str(row["node_id"]): str(row["node_label"])
            for _, row in common_nodes.iterrows()
        }
        for _, row in common_nodes.iterrows():
            axis = str(row["axis_role"])
            parent_label = labels_by_id.get(str(row.get("parent_node_id", "")), "")
            parent_code = _code_from_axis_label(parent_label)
            child_code = _code_from_axis_label(row.get("node_label"))
            if axis in components and parent_code and child_code and parent_code != child_code:
                children = components[axis].setdefault(parent_code, [])
                if child_code not in children:
                    children.append(child_code)

    if mappings_workbook_path.exists():
        import pandas as pd

        rollup_rules = pd.read_excel(mappings_workbook_path, sheet_name="esto_rollup_rules", dtype=object)
        if "include" in rollup_rules.columns:
            enabled = rollup_rules["include"].astype(str).str.strip().str.casefold().isin({"true", "1", "yes", "y"})
            rollup_rules = rollup_rules[enabled].copy()
        for rolled_label, group in rollup_rules.groupby("rolled_esto_flow", dropna=True):
            rolled_label = str(rolled_label).strip()
            rolled_code = _code_from_axis_label(rolled_label)
            member_codes: list[str] = []
            for member_label in group["input_esto_flow"].tolist():
                if pd.isna(member_label):
                    continue
                member_label = str(member_label or "").strip()
                if not member_label or member_label == rolled_label:
                    continue
                member_code = _code_from_axis_label(member_label)
                if member_code and member_code not in member_codes:
                    member_codes.append(member_code)
            if rolled_code and member_codes:
                components["flow"][rolled_code] = member_codes

    return {
        axis: {code: tuple(member_codes) for code, member_codes in axis_components.items()}
        for axis, axis_components in components.items()
    }


def resolve_json_synced_colors(
    base_by_axis: dict[str, dict[str, str]],
    exact_json_by_axis: dict[str, dict[str, str]],
    components_by_axis: dict[str, dict[str, tuple[str, ...]]],
    component_base_by_axis: dict[str, dict[str, str]] | None = None,
) -> dict[str, dict[str, str]]:
    """Resolve exact JSON colours, then OKLab component averages, then current colour.

    The final current-colour fallback is deliberate: opting into JSON sync can
    never replace a valid existing colour with a blank, ``NA``, or null value.
    """
    resolved: dict[str, dict[str, str]] = {"product": {}, "flow": {}}
    for axis in ("product", "flow"):
        base = {code: normalize_hex(color) for code, color in base_by_axis[axis].items()}
        exact = exact_json_by_axis[axis]
        components = components_by_axis.get(axis, {})
        component_bases = dict((component_base_by_axis or {}).get(axis, {}))

        def resolve(code: str, active: tuple[str, ...] = ()) -> str:
            if code in resolved[axis]:
                return resolved[axis][code]
            if code in exact:
                color = normalize_hex(exact[code])
            elif code in active:
                color = base.get(code, "")
            else:
                member_colors: list[str] = []
                for member_code in components.get(code, ()):
                    if member_code == code:
                        member_color = component_bases.get(code, base.get(code))
                    elif member_code in base or member_code in exact or member_code in components:
                        member_color = resolve(member_code, (*active, code))
                    else:
                        member_color = None
                    if member_color:
                        member_colors.append(member_color)
                color = average_oklab(member_colors) if member_colors else base.get(code, "")
            if color:
                resolved[axis][code] = color
            return color

        for code in base:
            resolve(code)
    return resolved


def _component_color_bases(
    payload: dict[str, object],
    components_by_axis: dict[str, dict[str, tuple[str, ...]]],
) -> dict[str, dict[str, str]]:
    """Preserve the uncombined colour for rollups that include their own code."""
    stored = dict(payload.get("_component_color_bases", {}))
    result: dict[str, dict[str, str]] = {"product": {}, "flow": {}}
    for axis in ("product", "flow"):
        result[axis] = {
            code: normalize_hex(color)
            for code, color in dict(stored.get(axis, {})).items()
        }
        axis_colors = dict(payload.get(axis, {}))
        for code, member_codes in components_by_axis.get(axis, {}).items():
            if code in member_codes and code in axis_colors:
                result[axis].setdefault(code, normalize_hex(axis_colors[code]))
    return result


# --- Workbook export ------------------------------------------------------

def _write_instructions(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "START HERE"
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "LEAP dashboard colour editor"
    sheet["A1"].fill = HEADER_FILL
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30

    instructions = [
        ("Choose colours", "Edit cells under Colour — EDIT. Use the paint bucket or type a hex colour such as #1F77B4."),
        ("JSON or manual", "TRUE follows colors.json. FALSE keeps your workbook colour."),
        ("Not directly in JSON", "If EXISTS_IN_JSON is FALSE and sync is TRUE, the colour is the OKLab average of its components. If no components exist, the current colour is kept."),
        ("Other categories", "These colours are always controlled by this workbook; they are not synced from JSON."),
        ("Keep it valid", "Do not rename tabs or categories, and do not add or delete rows. Save this same .xlsx file when finished."),
    ]
    sheet.append([])
    for heading, explanation in instructions:
        sheet.append([heading, explanation])
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 100
    for row in range(3, 8):
        sheet[f"A{row}"].font = Font(bold=True)
        sheet[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row].height = 34


def _write_colour_sheet(
    workbook: Workbook,
    sheet_name: str,
    rows: list[tuple[str, str, str]],
    sync_flags: dict[str, bool] | None = None,
    exists_in_json: dict[str, bool] | None = None,
) -> None:
    sheet = workbook.create_sheet(sheet_name)
    sheet.sheet_view.showGridLines = False
    # Column A is the stable machine identifier used during import. It stays
    # hidden so the reviewer sees one uncomplicated combined Category column.
    include_sync = sync_flags is not None
    headers = ["_internal_key", "Category", "Colour — EDIT"]
    if include_sync:
        headers.extend(["SYNC_WITH_JSON", "EXISTS_IN_JSON"])
    sheet.append(headers)
    for identifier, label, color in rows:
        row_number = sheet.max_row + 1
        values: list[object] = [identifier, label, color]
        if include_sync:
            values.append(str(bool(sync_flags.get(identifier, False))).upper())
            values.append(str(bool((exists_in_json or {}).get(identifier, False))).upper())
        sheet.append(values)
        color_cell = sheet.cell(row=row_number, column=3)
        color_cell.fill = PatternFill("solid", fgColor=color.lstrip("#"))
        color_cell.font = _font_for_fill(color)
        color_cell.alignment = Alignment(horizontal="center")
        sheet.cell(row=row_number, column=1).fill = LOCKED_FILL
        sheet.cell(row=row_number, column=2).fill = LOCKED_FILL
        sheet.cell(row=row_number, column=1).number_format = "@"

    header = sheet[1]
    for cell in header:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    header[2].fill = PatternFill("solid", fgColor="BF9000")
    if include_sync:
        header[3].fill = PatternFill("solid", fgColor="548235")
        header[4].fill = PatternFill("solid", fgColor="5B6573")
    sheet.row_dimensions[1].height = 34
    sheet.freeze_panes = "C2"
    # The Excel table below owns the filter. Adding a second worksheet-level
    # AutoFilter over the same cells creates conflicting OOXML that desktop
    # Excel repairs by removing the table.
    sheet.column_dimensions["A"].hidden = True
    sheet.column_dimensions["A"].width = 2
    sheet.column_dimensions["B"].width = 55
    sheet.column_dimensions["C"].width = 22
    if include_sync:
        sheet.column_dimensions["D"].width = 20
        sheet.column_dimensions["E"].width = 18
        for row_number in range(2, sheet.max_row + 1):
            sheet.cell(row=row_number, column=4).alignment = Alignment(horizontal="center")
            sheet.cell(row=row_number, column=5).alignment = Alignment(horizontal="center")
            sheet.cell(row=row_number, column=5).fill = LOCKED_FILL
        validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
        validation.error = "Choose TRUE to follow colors.json, or FALSE to keep the workbook colour."
        validation.errorTitle = "Choose TRUE or FALSE"
        validation.prompt = "TRUE = sync from JSON; FALSE = keep workbook colour"
        validation.promptTitle = "Colour source"
        validation.showErrorMessage = True
        validation.showInputMessage = True
        sheet.add_data_validation(validation)
        validation.add(f"D2:D{sheet.max_row}")
    if sheet.max_row > 1:
        sheet.conditional_formatting.add(
            f"C2:C{sheet.max_row}",
            FormulaRule(formula=["NOT(AND(LEFT(C2,1)=\"#\",LEN(C2)=7))"], fill=PatternFill("solid", fgColor="F4CCCC")),
        )
        table_name = re.sub(r"[^A-Za-z0-9]", "", sheet_name) + "Colours"
        last_column = "E" if include_sync else "C"
        table = Table(displayName=table_name, ref=f"A1:{last_column}{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)


def export_color_workbook(
    output_path: Path = DEFAULT_WORKBOOK_PATH,
    code_colors_path: Path = CODE_COLORS_PATH,
    common_rows_path: Path = DEFAULT_COMMON_ROWS_PATH,
    external_colors_path: Path | None = DEFAULT_EXTERNAL_COLORS_PATH,
    sync_flags: dict[str, dict[str, bool]] | None = None,
    axis_components: dict[str, dict[str, tuple[str, ...]]] | None = None,
) -> Path:
    """Create the workbook that can be sent directly to a colleague."""
    payload = json.loads(code_colors_path.read_text(encoding="utf-8"))
    source_config_hash = _config_hash(payload)
    external_source_hash = ""
    sync_colors: dict[str, dict[str, str]] = {"product": {}, "flow": {}}
    if external_colors_path is not None:
        sync_colors, external_source_hash = load_external_sync_colors(external_colors_path)
    base_by_axis = {
        axis: {code: normalize_hex(color) for code, color in dict(payload.get(axis, {})).items()}
        for axis in ("product", "flow")
    }
    components = axis_components if axis_components is not None else load_axis_color_components()
    component_bases = _component_color_bases(payload, components)
    resolved_sync_colors = resolve_json_synced_colors(
        base_by_axis,
        sync_colors,
        components,
        component_base_by_axis=component_bases,
    )
    labels = _load_axis_labels()
    workbook = Workbook()
    _write_instructions(workbook)
    workbook_metadata_rows: list[tuple[str, str, str, str, str, str]] = []

    for sheet_name, axis in SHEET_SPECS:
        mapping = dict(payload.get(axis, {}))
        axis_flags = dict((sync_flags or {}).get(axis, {}))
        rows = []
        for code, color in sorted(mapping.items()):
            category_name = str(labels.get(axis, {}).get(code, "")).strip()
            if not category_name or category_name == PLACEHOLDER_LABEL:
                continue
            follows_json = axis_flags.get(code, external_colors_path is not None)
            resolved = resolved_sync_colors[axis][code] if follows_json else color
            rows.append((code, f"{code} {category_name}", normalize_hex(resolved)))
            axis_flags[code] = follows_json
        exact_exists = {identifier: identifier in sync_colors[axis] for identifier, _label, _color in rows}
        _write_colour_sheet(
            workbook,
            sheet_name,
            rows,
            sync_flags=axis_flags,
            exists_in_json=exact_exists,
        )
        workbook_metadata_rows.extend(
            (
                sheet_name,
                identifier,
                label,
                color,
                str(axis_flags[identifier]).upper(),
                str(exact_exists[identifier]).upper(),
            )
            for identifier, label, color in rows
        )

    comparison_series = dict(dict(payload.get("plotting", {})).get("series", {}))
    other_rows = [
        (label, label, normalize_hex(color))
        for label, color in comparison_series.items()
    ]
    _write_colour_sheet(workbook, OTHER_SHEET_NAME, other_rows)
    workbook_metadata_rows.extend(
        (OTHER_SHEET_NAME, identifier, label, color, "", "")
        for identifier, label, color in other_rows
    )

    metadata = workbook.create_sheet("_metadata")
    metadata.sheet_state = "veryHidden"
    metadata.append(["schema_version", WORKBOOK_SCHEMA_VERSION])
    metadata.append(["source_config_sha256", source_config_hash])
    metadata.append(["source_config", str(code_colors_path)])
    metadata.append(["external_colors_sha256", external_source_hash])
    metadata.append(["external_colors_path", str(external_colors_path or "")])
    metadata.append([])
    metadata.append(["sheet", "key", "category", "current_color", "sync_with_json", "exists_in_json"])
    for metadata_row in workbook_metadata_rows:
        metadata.append(list(metadata_row))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"Wrote colleague colour workbook: {output_path}")
    return output_path


# --- Workbook import ------------------------------------------------------

def _chosen_colour(current_value: object, proposed_cell: object, location: str) -> str:
    current = normalize_hex(current_value)
    proposed_text = normalize_hex(getattr(proposed_cell, "value", None))
    proposed_fill = _cell_fill_hex(proposed_cell) or current
    text_changed = proposed_text != current
    fill_changed = proposed_fill != current
    if text_changed and fill_changed and proposed_text != proposed_fill:
        raise ValueError(
            f"{location}: typed colour {proposed_text} and fill colour {proposed_fill} disagree. "
            "Make them match or change only one."
        )
    return proposed_text if text_changed else proposed_fill


def _read_colour_sheet(
    workbook: object,
    sheet_name: str,
    workbook_metadata: dict[tuple[str, str], dict[str, str]],
    include_sync: bool,
) -> dict[str, dict[str, object]]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Required sheet {sheet_name!r} is missing or was renamed")
    sheet = workbook[sheet_name]
    expected_headers = ["Category", "Colour — EDIT"] + (
        ["SYNC_WITH_JSON", "EXISTS_IN_JSON"] if include_sync else []
    )
    actual_headers = [sheet.cell(row=1, column=column).value for column in range(2, 2 + len(expected_headers))]
    if actual_headers != expected_headers:
        raise ValueError(f"{sheet_name}: headings were changed; expected {expected_headers!r}")
    colors: dict[str, dict[str, object]] = {}
    for row_number in range(2, sheet.max_row + 1):
        identifier = str(sheet.cell(row=row_number, column=1).value or "").strip()
        if not identifier:
            if any(sheet.cell(row=row_number, column=column).value not in (None, "") for column in (2, 3)):
                raise ValueError(f"{sheet_name}!{row_number}: row has content but its protected identifier is missing")
            continue
        if identifier in colors:
            raise ValueError(f"{sheet_name}: duplicate identifier {identifier!r}")
        metadata_key = (sheet_name, identifier)
        if metadata_key not in workbook_metadata:
            raise ValueError(f"{sheet_name}: internal metadata is missing for {identifier!r}")
        metadata = workbook_metadata[metadata_key]
        expected_category = metadata["category"]
        actual_category = str(sheet.cell(row=row_number, column=2).value or "").strip()
        if actual_category != expected_category:
            raise ValueError(
                f"{sheet_name}!B{row_number}: category was changed; "
                f"expected {expected_category!r}, found {actual_category!r}"
            )
        current = normalize_hex(metadata["current_color"])
        chosen = _chosen_colour(
            current,
            sheet.cell(row=row_number, column=3),
            f"{sheet_name}!C{row_number}",
        )
        colors[identifier] = {
            "color": chosen,
            "changed": chosen != current,
            "current": current,
        }
        if include_sync:
            raw_sync = sheet.cell(row=row_number, column=4).value
            if isinstance(raw_sync, bool):
                follows_json = raw_sync
            elif str(raw_sync).strip().upper() in {"TRUE", "FALSE"}:
                follows_json = str(raw_sync).strip().upper() == "TRUE"
            else:
                raise ValueError(f"{sheet_name}!D{row_number}: choose TRUE or FALSE")
            colors[identifier]["sync_with_json"] = follows_json
            colors[identifier]["sync_changed"] = (
                str(follows_json).upper() != metadata.get("sync_with_json", "FALSE").upper()
            )
            actual_exists = str(sheet.cell(row=row_number, column=5).value or "").strip().upper()
            expected_exists = metadata.get("exists_in_json", "FALSE").upper()
            if actual_exists not in {"TRUE", "FALSE"} or actual_exists != expected_exists:
                raise ValueError(
                    f"{sheet_name}!E{row_number}: EXISTS_IN_JSON is automatic and must not be edited"
                )
    expected_keys = {
        key
        for (metadata_sheet, key), _details in workbook_metadata.items()
        if metadata_sheet == sheet_name
    }
    received_keys = set(colors)
    if received_keys != expected_keys:
        raise ValueError(
            f"{sheet_name}: category rows changed; "
            f"missing={sorted(expected_keys - received_keys)}, extra={sorted(received_keys - expected_keys)}"
        )
    return colors


def _read_workbook_metadata(workbook: object) -> dict[tuple[str, str], dict[str, str]]:
    """Read stable keys and comparison colours from the hidden metadata sheet."""
    sheet = workbook["_metadata"]
    metadata: dict[tuple[str, str], dict[str, str]] = {}
    header_row = next(
        (
            row_number
            for row_number in range(1, sheet.max_row + 1)
            if str(sheet.cell(row=row_number, column=1).value or "").strip() == "sheet"
        ),
        0,
    )
    if not header_row:
        raise ValueError("Internal workbook metadata table is missing")
    for row_number in range(header_row + 1, sheet.max_row + 1):
        sheet_name = str(sheet.cell(row=row_number, column=1).value or "").strip()
        identifier = str(sheet.cell(row=row_number, column=2).value or "").strip()
        if sheet_name and identifier:
            metadata_key = (sheet_name, identifier)
            if metadata_key in metadata:
                raise ValueError(f"Internal metadata contains duplicate row {metadata_key!r}")
            metadata[metadata_key] = {
                "category": str(sheet.cell(row=row_number, column=3).value or "").strip(),
                "current_color": normalize_hex(sheet.cell(row=row_number, column=4).value),
                "sync_with_json": str(sheet.cell(row=row_number, column=5).value or "").strip(),
                "exists_in_json": str(sheet.cell(row=row_number, column=6).value or "").strip(),
            }
    return metadata


def import_color_workbook(
    workbook_path: Path,
    code_colors_path: Path = CODE_COLORS_PATH,
    custom_colors_path: Path = CUSTOM_COLORS_PATH,
    common_rows_path: Path = DEFAULT_COMMON_ROWS_PATH,
    external_colors_path: Path | None = DEFAULT_EXTERNAL_COLORS_PATH,
    allow_stale_config: bool = False,
    refresh_workbook: bool = False,
    axis_components: dict[str, dict[str, tuple[str, ...]]] | None = None,
) -> Path:
    """Validate a returned workbook, save its scheme, and apply it to config."""
    workbook = load_workbook(workbook_path, data_only=False)
    if "_metadata" not in workbook.sheetnames:
        raise ValueError("This is not a dashboard colour workbook: hidden _metadata sheet is missing")
    schema_version = str(workbook["_metadata"]["B1"].value or "")
    if schema_version != WORKBOOK_SCHEMA_VERSION:
        raise ValueError(f"Unsupported workbook schema {schema_version!r}; expected {WORKBOOK_SCHEMA_VERSION!r}")
    workbook_metadata = _read_workbook_metadata(workbook)

    production_payload = json.loads(code_colors_path.read_text(encoding="utf-8"))
    workbook_source_hash = str(workbook["_metadata"]["B2"].value or "").strip()
    current_source_hash = _config_hash(production_payload)
    if workbook_source_hash != current_source_hash and not allow_stale_config:
        raise ValueError(
            "The dashboard colour configuration changed after this workbook was exported. "
            "Export a fresh workbook so newer categories or colours are not overwritten."
        )

    workbook_external_hash = str(workbook["_metadata"]["B4"].value or "").strip()
    sync_colors: dict[str, dict[str, str]] = {"product": {}, "flow": {}}
    current_external_hash = ""
    if external_colors_path is not None:
        sync_colors, current_external_hash = load_external_sync_colors(external_colors_path)

    custom_payload: dict[str, object] = {
        "_generated_by": "scripts/manage_dashboard_colors.py from a colleague-edited workbook",
        "_workbook": workbook_path.name,
        "product": {},
        "flow": {},
        "plotting": {"series": {}},
        "common_overrides": {"product": {}, "flow": {}},
        "_component_color_bases": {"product": {}, "flow": {}},
    }
    memberships = load_common_rollup_memberships(common_rows_path)
    retained_sync_flags: dict[str, dict[str, bool]] = {"product": {}, "flow": {}}
    workbook_changed = workbook_source_hash != current_source_hash or workbook_external_hash != current_external_hash
    entries_by_axis: dict[str, dict[str, dict[str, object]]] = {}
    proposed_base_by_axis: dict[str, dict[str, str]] = {}
    for sheet_name, axis in SHEET_SPECS:
        entries = _read_colour_sheet(workbook, sheet_name, workbook_metadata, include_sync=True)
        entries_by_axis[axis] = entries
        proposed_base = {
            code: normalize_hex(color)
            for code, color in dict(production_payload.get(axis, {})).items()
        }
        proposed_base.update({key: str(details["color"]) for key, details in entries.items()})
        proposed_base_by_axis[axis] = proposed_base
        for key, details in entries.items():
            retained_sync_flags[axis][key] = bool(details["sync_with_json"])
            workbook_changed = workbook_changed or bool(details["changed"]) or bool(details["sync_changed"])

    components = axis_components if axis_components is not None else load_axis_color_components()
    component_bases = _component_color_bases(production_payload, components)
    custom_payload["_component_color_bases"] = component_bases
    resolved_sync_colors = resolve_json_synced_colors(
        proposed_base_by_axis,
        sync_colors,
        components,
        component_base_by_axis=component_bases,
    )
    for _sheet_name, axis in SHEET_SPECS:
        entries = entries_by_axis[axis]
        edited_colors: dict[str, str] = {}
        for key, details in entries.items():
            follows_json = bool(details["sync_with_json"])
            if follows_json:
                edited_colors[key] = resolved_sync_colors[axis][key]
            else:
                edited_colors[key] = str(details["color"])
        merged_colors = dict(production_payload.get(axis, {}))
        merged_colors.update(edited_colors)
        custom_payload[axis] = merged_colors

    other_entries = _read_colour_sheet(workbook, OTHER_SHEET_NAME, workbook_metadata, include_sync=False)
    workbook_changed = workbook_changed or any(bool(details["changed"]) for details in other_entries.values())
    series_colors = {
        key: str(details["color"])
        for key, details in other_entries.items()
    }
    custom_payload["plotting"]["series"] = series_colors

    base_by_axis = {axis: dict(custom_payload[axis]) for axis in ("product", "flow")}
    resolved_common = build_common_rollup_colors(base_by_axis, memberships)
    production_payload["product"] = dict(custom_payload["product"])
    production_payload["flow"] = dict(custom_payload["flow"])
    production_payload.setdefault("plotting", {})["series"] = series_colors
    production_payload["common"] = resolved_common
    production_payload["_common_color_method"] = "equal-weight OKLab average of mapping-owned ESTO components"
    production_payload["_common_color_overrides"] = dict(custom_payload["common_overrides"])
    production_payload["_custom_color_source"] = custom_colors_path.name
    production_payload["_component_color_bases"] = component_bases

    custom_text = json.dumps(custom_payload, indent=2) + "\n"
    production_text = json.dumps(production_payload, indent=2) + "\n"
    config_changed = code_colors_path.read_text(encoding="utf-8") != production_text
    custom_changed = not custom_colors_path.exists() or custom_colors_path.read_text(encoding="utf-8") != custom_text
    if custom_changed:
        custom_colors_path.write_text(custom_text, encoding="utf-8")
    if config_changed:
        code_colors_path.write_text(production_text, encoding="utf-8")
        print(f"Applied workbook colours to: {code_colors_path}")
    if refresh_workbook and (workbook_changed or config_changed):
        try:
            export_color_workbook(
                output_path=workbook_path,
                code_colors_path=code_colors_path,
                common_rows_path=common_rows_path,
                external_colors_path=external_colors_path,
                sync_flags=retained_sync_flags,
                axis_components=components,
            )
        except PermissionError as exc:
            raise PermissionError(f"Close {workbook_path.name} in Excel so it can be synced.") from exc
    return custom_colors_path


def synchronize_dashboard_colors(
    workbook_path: Path = DEFAULT_WORKBOOK_PATH,
    code_colors_path: Path = CODE_COLORS_PATH,
    custom_colors_path: Path = CUSTOM_COLORS_PATH,
    common_rows_path: Path = DEFAULT_COMMON_ROWS_PATH,
    external_colors_path: Path = DEFAULT_EXTERNAL_COLORS_PATH,
    axis_components: dict[str, dict[str, tuple[str, ...]]] | None = None,
) -> Path:
    """Reconcile JSON-owned and workbook-owned colours before dashboard rendering."""
    load_external_sync_colors(external_colors_path)
    if not workbook_path.exists():
        export_color_workbook(
            output_path=workbook_path,
            code_colors_path=code_colors_path,
            common_rows_path=common_rows_path,
            external_colors_path=external_colors_path,
            axis_components=axis_components,
        )
    import_color_workbook(
        workbook_path=workbook_path,
        code_colors_path=code_colors_path,
        custom_colors_path=custom_colors_path,
        common_rows_path=common_rows_path,
        external_colors_path=external_colors_path,
        allow_stale_config=True,
        refresh_workbook=True,
        axis_components=axis_components,
    )
    print(f"Dashboard colours are synced: {workbook_path}")
    return workbook_path


# --- Frequently changed notebook-style controls --------------------------

EXPORT_WORKBOOK = False
IMPORT_WORKBOOK = False
WORKBOOK_PATH = DEFAULT_WORKBOOK_PATH
SYNC_EXTERNAL_COLORS = True
EXTERNAL_COLORS_PATH = DEFAULT_EXTERNAL_COLORS_PATH


#%%
if EXPORT_WORKBOOK:
    export_color_workbook(
        output_path=WORKBOOK_PATH,
        external_colors_path=EXTERNAL_COLORS_PATH if SYNC_EXTERNAL_COLORS else None,
    )

if IMPORT_WORKBOOK:
    import_color_workbook(
        workbook_path=WORKBOOK_PATH,
        external_colors_path=EXTERNAL_COLORS_PATH if SYNC_EXTERNAL_COLORS else None,
    )

#%%
